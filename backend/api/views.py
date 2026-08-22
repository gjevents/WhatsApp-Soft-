import csv
import json
import os
import asyncio
import time
from io import BytesIO

import requests
from django.db import transaction
from django.http import FileResponse, JsonResponse
from django.utils import timezone
from django.core.files.storage import default_storage
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Campaign, CampaignRecipient, Contact, ContactGroup, FailedMessage, MediaAsset, WhatsAppSettings
from .serializers import BulkContactUploadSerializer, CampaignSerializer, ContactGroupSerializer, ContactSerializer, FailedMessageSerializer

WHATSAPP_SERVICE_URL = os.getenv("WHATSAPP_SERVICE_URL", "http://localhost:3001")


def normalize_phone(raw):
    digits = "".join(ch for ch in str(raw) if ch.isdigit())
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def get_setting(key, default=""):
    setting, _ = WhatsAppSettings.objects.get_or_create(key=key, defaults={"value": default})
    return setting.value


def set_setting(key, value):
    obj, _ = WhatsAppSettings.objects.get_or_create(key=key)
    obj.value = str(value)
    obj.save()
    return obj


class ContactListCreateView(APIView):
    def get(self, request):
        contacts = Contact.objects.all().order_by("-created_at")
        serializer = ContactSerializer(contacts, many=True)
        return Response({"items": serializer.data, "count": contacts.count()})

    def post(self, request):
        data = request.data
        clean_number = normalize_phone(data.get("mobile") or data.get("phone"))
        if not clean_number:
            return Response({"error": "Mobile number is required."}, status=400)

        existing = Contact.objects.filter(mobile=clean_number).first()
        if existing:
            return Response({"error": "Contact with this mobile number already exists."}, status=400)

        payload = {"name": data.get("name", "").strip(), "mobile": clean_number, "consent_status": data.get("consent_status", "OPTED_IN")}
        serializer = ContactSerializer(data=payload)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

    def delete(self, request):
        contact_id = request.data.get("id")
        if not contact_id:
            return Response({"error": "Contact id required."}, status=400)
        Contact.objects.filter(id=contact_id).delete()
        return Response({"success": True})


class ContactDetailView(APIView):
    def put(self, request, pk):
        contact = Contact.objects.filter(pk=pk).first()
        if not contact:
            return Response({"error": "Contact not found."}, status=404)
        clean_number = normalize_phone(request.data.get("mobile") or contact.mobile)
        if not clean_number:
            return Response({"error": "Mobile number is required."}, status=400)
        if Contact.objects.filter(mobile=clean_number).exclude(pk=pk).exists():
            return Response({"error": "Contact with this mobile number already exists."}, status=400)
        contact.name = str(request.data.get("name", contact.name)).strip()
        contact.mobile = clean_number
        if "consent_status" in request.data:
            contact.consent_status = request.data["consent_status"]
        serializer = ContactSerializer(contact, data={
            "name": contact.name,
            "mobile": contact.mobile,
            "consent_status": contact.consent_status,
        }, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, pk):
        deleted, _ = Contact.objects.filter(pk=pk).delete()
        if not deleted:
            return Response({"error": "Contact not found."}, status=404)
        return Response({"success": True})


class ContactImportView(APIView):
    def post(self, request):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"error": "CSV or Excel file is required."}, status=400)

        filename = file_obj.name.lower()
        if filename.endswith(".csv"):
            text = file_obj.read().decode("utf-8-sig")
            rows = list(csv.DictReader(text.splitlines()))
        elif filename.endswith((".xlsx", ".xls")):
            workbook = load_workbook(BytesIO(file_obj.read()), data_only=True)
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                if not any(cell is not None for cell in row):
                    continue
                rows.append(dict(zip(["name", "mobile", "consent_status"], row[:3])))
        else:
            return Response({"error": "Unsupported file type. Please upload CSV or Excel."}, status=400)

        created = 0
        duplicates = 0
        invalid = 0
        for row in rows:
            if not row:
                continue
            name = str(row.get("name") or row.get("Name") or "").strip()
            mobile = normalize_phone(row.get("mobile") or row.get("phone") or row.get("Mobile") or row.get("Phone") or "")
            consent = str(row.get("consent_status") or row.get("consentStatus") or row.get("Consent") or "OPTED_IN").upper()
            if not mobile:
                invalid += 1
                continue
            if Contact.objects.filter(mobile=mobile).exists():
                duplicates += 1
                continue
            Contact.objects.create(name=name or "Unknown Contact", mobile=mobile, consent_status=consent if consent in {"OPTED_IN", "PENDING", "OPTED_OUT"} else "OPTED_IN")
            created += 1

        return Response({
            "success": True,
            "created": created,
            "duplicates": duplicates,
            "invalid": invalid,
            "count": Contact.objects.count(),
        })


class CampaignListCreateView(APIView):
    def get(self, request):
        campaigns = Campaign.objects.all().order_by("-created_at")
        serializer = CampaignSerializer(campaigns, many=True)
        return Response({"items": serializer.data, "count": campaigns.count()})

    def post(self, request):
        payload = {
            "name": request.data.get("name", "").strip(),
            "message": request.data.get("message", "").strip(),
            "delay_seconds": int(request.data.get("delay_seconds") or 5),
            "status": "DRAFT",
        }
        campaign = Campaign.objects.create(**payload)

        media_ids = request.data.get("media_ids") or []
        if isinstance(media_ids, str):
            media_ids = json.loads(media_ids)
        campaign.media.set(MediaAsset.objects.filter(id__in=media_ids))

        contacts = request.data.get("selected_contact_ids") or []
        group_id = request.data.get("group_id")
        if group_id:
            group = ContactGroup.objects.filter(id=group_id).first()
            if not group:
                return Response({"error": "Selected group was not found."}, status=400)
            contacts = list(group.contacts.filter(consent_status="OPTED_IN").values_list("id", flat=True))
            if not contacts:
                return Response({"error": "Selected group has no opted-in contacts."}, status=400)
        if isinstance(contacts, str):
            contacts = json.loads(contacts)

        recipients = []
        for contact_id in contacts:
            contact = Contact.objects.filter(id=contact_id, consent_status="OPTED_IN").first()
            if not contact:
                continue
            recipients.append(CampaignRecipient(campaign=campaign, contact=contact, phone=contact.mobile, status="PENDING"))
        CampaignRecipient.objects.bulk_create(recipients)
        campaign.total_recipients = len(recipients)
        campaign.pending_count = len(recipients)
        campaign.save()
        serializer = CampaignSerializer(campaign)
        return Response(serializer.data, status=201)


class CampaignDetailView(APIView):
    def get(self, request, pk):
        campaign = Campaign.objects.get(pk=pk)
        return Response(CampaignSerializer(campaign).data)


class CampaignSendView(APIView):
    def post(self, request, pk):
        campaign = Campaign.objects.get(pk=pk)
        campaign.status = "QUEUED"
        campaign.save()
        
        # Start sending messages asynchronously
        self._send_campaign_messages(campaign)
        
        return Response({"success": True, "campaign_id": campaign.id, "status": campaign.status})
    
    def _send_campaign_messages(self, campaign):
        """Send messages to all pending recipients with delays"""
        import threading
        thread = threading.Thread(target=self._send_messages_thread, args=(campaign,))
        thread.daemon = True
        thread.start()
    
    def _send_messages_thread(self, campaign):
        """Background thread to send messages"""
        try:
            recipients = CampaignRecipient.objects.filter(campaign=campaign, status="PENDING").order_by("id")
            
            for recipient in recipients:
                if campaign.status == "STOPPED":
                    break
                
                if campaign.status == "PAUSED":
                    # Wait while paused
                    while campaign.status == "PAUSED":
                        campaign.refresh_from_db()
                        time.sleep(1)
                    if campaign.status == "STOPPED":
                        break
                
                # Send message via WhatsApp service
                try:
                    media_assets = list(campaign.media.all())
                    endpoint = "/send-media" if media_assets else "/send"
                    payload = {"phone": recipient.phone, "message": campaign.message}
                    if media_assets:
                        payload.update({
                            "mediaUrls": [request_media_url(asset) for asset in media_assets],
                            "mediaType": media_assets[0].file_type,
                        })
                    response = requests.post(f"{WHATSAPP_SERVICE_URL}{endpoint}", json=payload, timeout=60)

                    if response.status_code == 200:
                        recipient.status = "SENT"
                        recipient.attempts += 1
                        recipient.last_attempt_at = timezone.now()
                        recipient.save()
                        campaign.sent_count += 1
                        campaign.pending_count -= 1
                        campaign.save()
                    else:
                        error_data = response.json() if response.headers.get("Content-Type", "").startswith("application/json") else {"error": response.text}
                        raise Exception(error_data.get("error", f"HTTP {response.status_code}"))

                except Exception as exc:
                    recipient.attempts += 1
                    recipient.last_attempt_at = timezone.now()

                    if recipient.attempts >= 3:
                        recipient.status = "FAILED"
                        campaign.failed_count += 1
                        campaign.pending_count -= 1
                    else:
                        recipient.status = "PENDING"

                    FailedMessage.objects.create(
                        campaign=campaign,
                        contact=recipient.contact,
                        phone=recipient.phone,
                        error=str(exc),
                        attempts=recipient.attempts,
                        status=recipient.status,
                    )

                    recipient.save()
                    campaign.save()
                
                # Delay between messages
                if campaign.delay_seconds and recipient != list(recipients)[-1]:
                    time.sleep(campaign.delay_seconds)
            
            # Mark campaign as completed if all sent or stopped
            if campaign.pending_count == 0:
                campaign.status = "COMPLETED"
            elif campaign.status == "STOPPED":
                campaign.status = "STOPPED"
            
            campaign.save()
        
        except Exception as exc:
            campaign.status = "FAILED"
            campaign.save()
            print(f"Campaign {campaign.id} send failed: {exc}")


class CampaignPauseView(APIView):
    def post(self, request, pk):
        campaign = Campaign.objects.get(pk=pk)
        campaign.status = "PAUSED"
        campaign.save()
        return Response({"success": True, "status": campaign.status})


class CampaignResumeView(APIView):
    def post(self, request, pk):
        campaign = Campaign.objects.get(pk=pk)
        campaign.status = "QUEUED"
        campaign.save()
        return Response({"success": True, "status": campaign.status})


class CampaignStopView(APIView):
    def post(self, request, pk):
        campaign = Campaign.objects.get(pk=pk)
        campaign.status = "STOPPED"
        campaign.save()
        return Response({"success": True, "status": campaign.status})


class CampaignRetryFailedView(APIView):
    def post(self, request, pk):
        campaign = Campaign.objects.get(pk=pk)
        failed = CampaignRecipient.objects.filter(campaign=campaign, status="FAILED")
        updated = 0
        for recipient in failed:
            if recipient.attempts >= 3:
                continue
            recipient.status = "PENDING"
            recipient.attempts += 1
            recipient.save()
            updated += 1
        return Response({"success": True, "updated": updated})


class MediaUploadView(APIView):
    def get(self, request):
        assets = MediaAsset.objects.all().order_by("-created_at")
        return Response({
            "items": [
                {
                    "id": asset.id,
                    "file_name": asset.file_name,
                    "original_name": asset.original_name,
                    "file_type": asset.file_type,
                    "file_size": asset.file_size,
                    "storage_path": asset.storage_path,
                    "url": request.build_absolute_uri(f"/media/{asset.storage_path.replace(os.sep, '/') }"),
                    "download_url": request.build_absolute_uri(f"/api/media/{asset.id}/download"),
                    "created_at": asset.created_at,
                }
                for asset in assets
            ],
            "count": assets.count(),
        })

    def post(self, request):
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"error": "No file uploaded"}, status=400)
        allowed = {"jpg", "jpeg", "png", "webp", "pdf", "mp4", "mov", "doc", "docx", "xls", "xlsx", "ppt", "pptx"}
        ext = os.path.splitext(uploaded.name)[1].lower().replace(".", "")
        if ext in {"exe", "bat", "cmd", "msi", "dll", "scr"} or ext not in allowed:
            return Response({"error": "Unsupported or executable file type"}, status=400)

        path = os.path.join("uploads", uploaded.name)
        with default_storage.open(path, "wb+") as destination:
            for chunk in uploaded.chunks():
                destination.write(chunk)

        media = MediaAsset.objects.create(
            file_name=uploaded.name,
            original_name=uploaded.name,
            file_type=ext,
            file_size=uploaded.size,
            storage_path=path,
        )
        return Response({"success": True, "media": {
            "id": media.id,
            "file_name": media.file_name,
            "original_name": media.original_name,
            "file_type": media.file_type,
            "file_size": media.file_size,
            "storage_path": media.storage_path,
            "url": request.build_absolute_uri(f"/media/{media.storage_path.replace(os.sep, '/') }"),
            "download_url": request.build_absolute_uri(f"/api/media/{media.id}/download"),
        }})


class MediaDownloadView(APIView):
    def get(self, request, pk):
        try:
            media = MediaAsset.objects.get(pk=pk)
            file_handle = default_storage.open(media.storage_path, "rb")
        except (MediaAsset.DoesNotExist, FileNotFoundError):
            return Response({"error": "Media file not found"}, status=404)
        return FileResponse(file_handle, as_attachment=True, filename=media.original_name)


def request_media_url(media):
    base_url = os.getenv("BACKEND_PUBLIC_URL", "http://localhost:8000")
    return f"{base_url}/media/{media.storage_path.replace(os.sep, '/')}"


class GroupListCreateView(APIView):
    def get(self, request):
        groups = ContactGroup.objects.prefetch_related("contacts").order_by("name")
        return Response({"items": ContactGroupSerializer(groups, many=True).data})

    def post(self, request):
        name = str(request.data.get("name", "")).strip()
        contact_ids = request.data.get("contact_ids") or []
        try:
            max_members = int(request.data.get("max_members", 60))
        except (TypeError, ValueError):
            return Response({"error": "Group contact limit must be a whole number."}, status=400)
        if not name:
            return Response({"error": "Group name is required."}, status=400)
        if max_members < 1:
            return Response({"error": "Group contact limit must be at least 1."}, status=400)
        if len(contact_ids) > max_members:
            return Response({"error": f"This group allows up to {max_members} contacts."}, status=400)
        if ContactGroup.objects.filter(name__iexact=name).exists():
            return Response({"error": "A group with this name already exists."}, status=400)
        assigned_contacts = Contact.objects.filter(id__in=contact_ids, groups__isnull=False).distinct()
        if assigned_contacts.exists():
            assigned_names = ", ".join(assigned_contacts.values_list("name", flat=True)[:5])
            return Response({"error": f"These contacts are already in another group: {assigned_names}."}, status=400)
        contacts = Contact.objects.filter(id__in=contact_ids, consent_status="OPTED_IN", groups__isnull=True).distinct()
        group = ContactGroup.objects.create(name=name, max_members=max_members)
        group.contacts.set(contacts)
        return Response(ContactGroupSerializer(group).data, status=201)


class GroupDetailView(APIView):
    def get(self, request, pk):
        return Response(ContactGroupSerializer(ContactGroup.objects.prefetch_related("contacts").get(pk=pk)).data)

    def put(self, request, pk):
        group = ContactGroup.objects.get(pk=pk)
        next_name = str(request.data.get("name", group.name)).strip()
        if not next_name:
            return Response({"error": "Group name is required."}, status=400)
        if ContactGroup.objects.filter(name__iexact=next_name).exclude(pk=group.pk).exists():
            return Response({"error": "A group with this name already exists."}, status=400)
        try:
            max_members = int(request.data.get("max_members", group.max_members))
        except (TypeError, ValueError):
            return Response({"error": "Group contact limit must be a whole number."}, status=400)
        if max_members < 1:
            return Response({"error": "Group contact limit must be at least 1."}, status=400)
        group.name = next_name
        contact_ids = request.data.get("contact_ids", list(group.contacts.values_list("id", flat=True)))
        if len(contact_ids) > max_members:
            return Response({"error": f"This group allows up to {max_members} contacts."}, status=400)
        contacts = Contact.objects.filter(id__in=contact_ids, consent_status="OPTED_IN").distinct()
        if len(contact_ids) != contacts.count() or contacts.filter(groups__isnull=False).exclude(groups=group).exists():
            return Response({"error": "One or more contacts are already assigned to another group or are not opted in."}, status=400)
        group.max_members = max_members
        group.save()
        group.contacts.set(contacts)
        return Response(ContactGroupSerializer(group).data)

    def delete(self, request, pk):
        ContactGroup.objects.filter(pk=pk).delete()
        return Response({"success": True})


class SettingsView(APIView):
    def get(self, request):
        return Response({
            "connected_phone": get_setting("connected_phone", ""),
            "default_delay_seconds": int(get_setting("default_delay_seconds", "5") or 5),
        })

    def put(self, request):
        for key in ("connected_phone", "default_delay_seconds"):
            if key in request.data:
                set_setting(key, request.data[key])
        return self.get(request)


class ReportsView(APIView):
    def get(self, request):
        campaigns = Campaign.objects.all()
        data = {
            "campaigns": [
                {
                    "id": c.id,
                    "name": c.name,
                    "status": c.status,
                    "total": c.total_recipients,
                    "sent": c.sent_count,
                    "failed": c.failed_count,
                    "pending": c.pending_count,
                    "success_rate": round((c.sent_count / c.total_recipients) * 100, 2) if c.total_recipients else 0,
                    "date": c.created_at.isoformat(),
                }
                for c in campaigns
            ]
        }
        return Response(data)


class DashboardView(APIView):
    def get(self, request):
        try:
            status_response = requests.get(f"{WHATSAPP_SERVICE_URL}/state", timeout=5)
            whatsapp_status = status_response.json().get("status", "DISCONNECTED") if status_response.ok else "DISCONNECTED"
        except Exception:
            whatsapp_status = "DISCONNECTED"

        data = {
            "total_contacts": Contact.objects.count(),
            "total_campaigns": Campaign.objects.count(),
            "messages_sent": sum(c.sent_count for c in Campaign.objects.all()),
            "messages_failed": sum(c.failed_count for c in Campaign.objects.all()),
            "messages_pending": sum(c.pending_count for c in Campaign.objects.all()),
            "whatsapp_status": whatsapp_status,
            "contacts_by_consent": {
                "OPTED_IN": Contact.objects.filter(consent_status="OPTED_IN").count(),
                "PENDING": Contact.objects.filter(consent_status="PENDING").count(),
                "OPTED_OUT": Contact.objects.filter(consent_status="OPTED_OUT").count(),
            },
        }
        return Response(data)


class FailedMessagesView(APIView):
    def get(self, request):
        messages = FailedMessage.objects.select_related("contact").all().order_by("-created_at")
        serializer = FailedMessageSerializer(messages, many=True)
        return Response({"items": serializer.data, "count": messages.count()})


class WhatsAppStatusView(APIView):
    def get(self, request):
        try:
            response = requests.get(f"{WHATSAPP_SERVICE_URL}/state", timeout=5)
            if response.ok:
                return Response(response.json())
        except Exception:
            pass
        return Response({"status": "DISCONNECTED", "qr": None, "message": "WhatsApp service unavailable"})


class WhatsAppConnectView(APIView):
    def post(self, request):
        try:
            expected_phone = request.data.get("phone") or get_setting("connected_phone", "")
            response = requests.post(f"{WHATSAPP_SERVICE_URL}/connect", json={"expectedPhone": expected_phone}, timeout=20)
            return Response(response.json(), status=response.status_code)
        except Exception as exc:
            return Response({
                "status": "DISCONNECTED",
                "message": "WhatsApp service is not running. Restart with `python manage.py runserver`.",
                "detail": str(exc),
            }, status=503)


class WhatsAppDisconnectView(APIView):
    def post(self, request):
        try:
            response = requests.post(f"{WHATSAPP_SERVICE_URL}/disconnect", timeout=20)
            return Response(response.json(), status=response.status_code)
        except Exception as exc:
            return Response({
                "status": "DISCONNECTED",
                "qr": None,
                "message": "WhatsApp service unavailable; it is already disconnected.",
                "detail": str(exc),
            })
